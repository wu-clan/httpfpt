#!/usr/bin/env python3
# -*- coding: utf-8 -*-
import os
import shutil
from datetime import datetime
from typing import Any, Dict, List

import xlrd
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, Color
from openpyxl.styles.colors import COLOR_INDEX

from fastpt.common.log import log
from fastpt.core.get_conf import TESTER_NAME
from fastpt.core.path_conf import EXCEL_DATA_PATH, EXCEL_REPORT_PATH


def read_excel(filepath: str = EXCEL_DATA_PATH, *, filename: str, sheet: str = 'Sheet1') -> List[Dict[str, Any]]:
    """
    读取 xlsx 文件

    :param filepath: 文件路径
    :param filename: 文件名
    :param sheet: 工作表
    :return:
    """
    file = os.path.join(filepath, filename)
    data = xlrd.open_workbook(file)
    table = data.sheet_by_name(sheet)
    # 获取总行,列数
    rows = table.nrows
    cols = table.ncols
    if rows > 1:
        # 获取第一行内容, 通常为列说明
        keys = table.row_values(0)
        data_list = []
        # 获取文档剩下所有内容
        for col in range(1, rows):
            values = table.row_values(col)
            # key, value组合为字典
            data = dict(zip(keys, values))
            # 数据整理
            for value in data:
                if isinstance(data[value], str):
                    data[value] = data[value].replace(' ', '').replace('\n', '')
                if data[value] == '':
                    data[value] = None
            data_list.append(data)
        return data_list
    else:
        log.warning('数据表格没有数据!')
        raise ValueError('数据表格没有数据! 请检查数据文件内容是否正确!')


def write_excel_report(datafile='APITestCaseTEMPLATE.xlsx',
                       filename: str = f'APITestResult_{datetime.now().strftime("%Y-%m-%d %H_%M_%S")}.xlsx', *,
                       row_num: int, status: str):
    """
    写入 excel 测试报告

    :param datafile: excel测试数据文件名
    :param filename: 文件名
    :param row_num:数据所在行数
    :param status: 测试结果: 'PASS' or 'FAIL'
    :return
    """
    if status not in ('PASS', 'FAIL', 'SKIP'):
        raise ValueError('excel测试报告结果用力状态只允许"PASS","FAIL"或"SKIP"')
    if not os.path.exists(EXCEL_REPORT_PATH):
        os.makedirs(EXCEL_REPORT_PATH)
    _datafile = os.path.join(EXCEL_DATA_PATH, datafile)
    _report_file = os.path.join(EXCEL_REPORT_PATH, filename)
    # copy测试数据为报告文件基础
    shutil.copyfile(_datafile, _report_file)
    wb = load_workbook(_report_file)
    ws = wb.active
    font_green = Font(name='宋体', color=Color(rgb=COLOR_INDEX[3]), bold=True)
    font_red = Font(name='宋体', color=Color(rgb=COLOR_INDEX[2]), bold=True)
    font_yellow = Font(name='宋体', color=Color(rgb=COLOR_INDEX[5]), bold=True)
    font_black = Font(name='宋体', color=Color(), bold=True)
    # 所在行,列
    wl = "L" + str(row_num)
    wm = "M" + str(row_num)
    # todo 新版写入待更新
    # 用力状态
    if status == "PASS":
        ws.cell(row_num, 12, status)
        ws[wl].font = font_green
    elif status == "FAIL":
        ws.cell(row_num, 12, status)
        ws[wl].font = font_red
    elif status == "SKIP":
        ws.cell(row_num, 12, status)
        ws[wl].font = font_yellow
    # 测试员
    ws.cell(row_num, 13, TESTER_NAME)
    ws[wm].font = font_black
    # 文件内容格式
    align = Alignment(horizontal='center', vertical='center')
    ws[wl].alignment = ws[wm].alignment = align
    try:
        wb.save(_report_file)
    except Exception as e:
        log.error(f'写入excel测试报告失败 \n {e}')
        raise e
    else:
        if status == 'PASS':
            log.success('test result: ----> {}', status)
        elif status == 'FAIL':
            log.error('test result: ----> {}', status)
        elif status == 'SKIP':
            log.warning('test result: ----> {}', status)
        log.success('写入excel测试报告成功')


def get_excel_row(data: dict):
    """
    写入 excel 文件的行数

    :param data:
    :return:
    """
    row_num = int(data['case_id'].split("_")[2]) + 1
    return row_num
