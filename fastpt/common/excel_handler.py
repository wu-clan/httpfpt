#!/usr/bin/env python3
# -*- coding: utf-8 -*-
import os
import shutil
from datetime import datetime
from typing import Any, Dict, List, Union, Optional

import xlrd
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment

from fastpt.common.log import log
from fastpt.core.get_conf import TESTER_NAME
from fastpt.core.path_conf import EXCEL_DATA_PATH, EXCEL_REPORT_PATH


def read_excel(
        filepath: str = EXCEL_DATA_PATH,
        *,
        filename: str,
        sheet: str = 'Sheet1'
) -> List[Dict[str, Optional[Any]]]:
    """
    读取 xlsx 文件

    :param filepath: 文件路径
    :param filename: 文件名
    :param sheet: 工作表
    :return:
    """
    file = os.path.join(filepath, filename)
    data = xlrd.open_workbook(file)
    table = data.sheet_by_name(sheet)
    # 获取总行,列数
    rows = table.nrows
    cols = table.ncols
    if rows > 1:
        # 获取第一行内容, 通常为列说明
        keys = table.row_values(0)
        data_list = []
        # 获取文档剩下所有内容
        for col in range(1, rows):
            values = table.row_values(col)
            # key, value组合为字典
            data = dict(zip(keys, values))
            # 数据整理
            for value in data:
                # 替换空字符串为 None, 保证与 yaml 数据返回格式一致
                if data[value] == '':
                    data[value] = None
            data_list.append(data)
        return data_list
    else:
        log.warning(f'数据表格 {filename} 没有数据!')
        raise ValueError(f'数据表格 {filename} 没有数据! 请检查数据文件内容是否正确!')


def write_excel_report(
        datafile='APITestCaseTEMPLATE.xlsx',
        filename: str = f'APITestResult_{datetime.now().strftime("%Y-%m-%d %H_%M_%S")}.xlsx',
        *,
        row_num: int,
        status: str,
        extension: Union[str, None] = None
) -> None:
    """
    写入 excel 测试报告

    :param datafile: excel测试数据文件名
    :param filename: 文件名
    :param row_num: 数据写入行
    :param status: 测试结果: PASS / FAIL
    :param extension: 测试结果扩展信息，如果有，则覆盖默认写入的结果
    :return
    """
    status_upper = status.upper()
    if status_upper not in ('PASS', 'FAIL'):
        raise ValueError('excel测试报告结果用力状态只允许"PASS","FAIL')
    if not os.path.exists(EXCEL_REPORT_PATH):
        os.makedirs(EXCEL_REPORT_PATH)
    data_file = os.path.join(EXCEL_DATA_PATH, datafile)
    report_file = os.path.join(EXCEL_REPORT_PATH, filename)
    if not os.path.exists(report_file):
        # copy测试数据为报告文件基础
        shutil.copyfile(data_file, report_file)
    wb = load_workbook(report_file)
    wa = wb.active
    # 字体颜色
    green = Font(name='Consolas', color='99CC00', bold=True)
    red = Font(name='Consolas', color='FF0000', bold=True)
    black = Font(name='Consolas', color='000000', bold=True)
    # 文件内容格式
    align = Alignment(horizontal='left', vertical='center')
    # 所在指定单元格: 列字母 + 所在行
    result_title_box = 'O1'  # 字母 O，不是数字 0
    result_box = "O" + str(row_num)
    tester_title_box = "P1"
    tester_box = "P2"
    # 结果写入列
    result_col = 15
    tester_col = result_col + 1
    # 写入测试结果
    if not wa[result_title_box].value:
        wa.cell(1, result_col, 'result')
        wa[result_title_box].font = black
    if extension:
        wa.cell(row_num, result_col, extension)
    else:
        wa.cell(row_num, result_col, status_upper)
    if status_upper == "PASS":
        wa[result_box].font = green
    elif status_upper == "FAIL":
        wa[result_box].font = red
    # 写入测试员
    if not wa[tester_box].value:
        wa.cell(1, tester_col, 'tester')
        wa[tester_title_box].font = black
        wa.cell(2, tester_col, TESTER_NAME)
        wa[tester_box].font = black
    # 修改写入单元格样式
    wa[result_box].alignment = wa[tester_box].alignment = align
    try:
        wb.save(report_file)
    except Exception as e:
        log.error(f'写入 {filename} 测试报告失败: {e}')
        raise e
    else:
        log.success(f'写入 {filename} 测试报告成功')


def get_excel_row(data: dict) -> int:
    """
    写入 excel 文件的行数

    :param data:
    :return:
    """
    row_num = int(data['case_id'].split("_")[2]) + 1
    return row_num
